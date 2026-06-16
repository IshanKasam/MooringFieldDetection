"""Export fields and prospects to Excel and CSV."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from mooring_fields.database import (
    fields_export_rows,
    get_connection,
    init_db,
    prospects_export_rows,
    supply_chain_export_rows,
)

FIELDS_COLUMNS = [
    "field_id",
    "scan_id",
    "latitude",
    "longitude",
    "boat_count",
    "mean_confidence",
    "location_name",
    "country",
    "detection_weights",
    "detection_date",
    "enriched_place_name",
    "enrichment_status",
    "prospect_id",
    "needs_review",
]

PROSPECTS_COLUMNS = [
    "prospect_id",
    "canonical_business_name",
    "phone",
    "email",
    "website",
    "address",
    "operator_type",
    "harbor_name",
    "research_summary",
    "supply_chain_summary",
    "confidence",
    "sources",
    "field_ids",
    "field_count",
    "needs_review",
    "approved",
    "last_enriched",
]

SUPPLY_CHAIN_COLUMNS = [
    "prospect_id",
    "mooring_company",
    "harbor_name",
    "field_ids",
    "supplier_or_manufacturer",
    "component_types",
    "evidence",
    "confidence_level",
    "confirmation_status",
    "notable_brands",
    "company_overall_confidence",
]


def _safe_cell(value: Any) -> Any:
    """Normalize cell values for CSV/Excel (no control chars, bounded length)."""
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # Excel cell limit ~32k chars
    if len(text) > 32000:
        text = text[:31997] + "..."
    return text


def _sanitize_rows(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        out.append({col: _safe_cell(row.get(col)) for col in columns})
    return out


def _filter_rows(
    fields: list[dict[str, Any]],
    prospects: list[dict[str, Any]],
    *,
    min_boat_count: int = 0,
    min_confidence: float = 0.0,
    only_approved: bool = False,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    approved_ids = {
        p["prospect_id"]
        for p in prospects
        if p.get("approved") and (not only_approved or p.get("approved"))
    }
    if only_approved:
        prospects = [p for p in prospects if p.get("approved")]
    else:
        prospects = [
            p
            for p in prospects
            if float(p.get("confidence") or 0) >= min_confidence
        ]
    fields = [
        f
        for f in fields
        if int(f.get("boat_count") or 0) >= min_boat_count
        and (
            not only_approved
            or f.get("prospect_id") in approved_ids
            or f.get("prospect_id") is None
        )
    ]
    return fields, prospects


def export_csv_pair(
    output_dir: Path,
    fields: list[dict[str, Any]],
    prospects: list[dict[str, Any]],
    supply_chain: list[dict[str, Any]] | None = None,
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    fields_path = output_dir / "fields_export.csv"
    prospects_path = output_dir / "prospects_export.csv"
    supply_path = output_dir / "supply_chain_export.csv"

    fields = _sanitize_rows(fields, FIELDS_COLUMNS)
    prospects = _sanitize_rows(prospects, PROSPECTS_COLUMNS)

    with fields_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS_COLUMNS, extrasaction="ignore")
        w.writeheader()
        w.writerows(fields)
    with prospects_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=PROSPECTS_COLUMNS, extrasaction="ignore")
        w.writeheader()
        w.writerows(prospects)

    result = {"fields_csv": str(fields_path), "prospects_csv": str(prospects_path)}
    if supply_chain is not None:
        supply_chain = _sanitize_rows(supply_chain, SUPPLY_CHAIN_COLUMNS)
        with supply_path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=SUPPLY_CHAIN_COLUMNS, extrasaction="ignore")
            w.writeheader()
            w.writerows(supply_chain)
        result["supply_chain_csv"] = str(supply_path)
    return result


def export_excel(
    xlsx_path: Path,
    fields: list[dict[str, Any]],
    prospects: list[dict[str, Any]],
    supply_chain: list[dict[str, Any]] | None = None,
) -> str:
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        ) from e

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    fields = _sanitize_rows(fields, FIELDS_COLUMNS)
    prospects = _sanitize_rows(prospects, PROSPECTS_COLUMNS)

    wb = Workbook()
    ws_fields = wb.active
    ws_fields.title = "Fields"
    ws_fields.append(FIELDS_COLUMNS)
    for row in fields:
        ws_fields.append([row.get(c) for c in FIELDS_COLUMNS])

    ws_prospects = wb.create_sheet("Prospects")
    ws_prospects.append(PROSPECTS_COLUMNS)
    for row in prospects:
        ws_prospects.append([row.get(c) for c in PROSPECTS_COLUMNS])

    if supply_chain is not None:
        supply_chain = _sanitize_rows(supply_chain, SUPPLY_CHAIN_COLUMNS)
        ws_supply = wb.create_sheet("Supply_Chain")
        ws_supply.append(SUPPLY_CHAIN_COLUMNS)
        for row in supply_chain:
            ws_supply.append([row.get(c) for c in SUPPLY_CHAIN_COLUMNS])

    wb.save(xlsx_path)
    return str(xlsx_path)


def export_prospects(
    xlsx_path: Path,
    *,
    db_path: Path | None = None,
    min_boat_count: int = 0,
    min_confidence: float = 0.0,
    only_approved: bool = False,
    also_csv: bool = True,
) -> dict[str, str]:
    conn = get_connection(db_path)
    try:
        init_db(conn)
        fields = fields_export_rows(conn)
        prospects = prospects_export_rows(conn)
        supply_chain = supply_chain_export_rows(conn)
    finally:
        conn.close()

    fields, prospects = _filter_rows(
        fields,
        prospects,
        min_boat_count=min_boat_count,
        min_confidence=min_confidence,
        only_approved=only_approved,
    )
    result = {
        "xlsx": export_excel(xlsx_path, fields, prospects, supply_chain),
        "supply_chain_rows": str(len(supply_chain)),
    }
    if also_csv:
        result.update(export_csv_pair(xlsx_path.parent, fields, prospects, supply_chain))
    return result
